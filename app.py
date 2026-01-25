# -*- coding: utf-8 -*-
"""
Peer-distance research webapp (Flask + PostgreSQL)

Design intent:
- This app does NOT diagnose/evaluate students.
- It collects perceived relationship structure via spatial placements and summarizes structure via
  distance-based analysis, MDS, and (optional) K-means.

Operational notes:
- PostgreSQL is assumed (JSONB used).
- init_db() includes a minimal schema versioning + migrations skeleton.
"""

from __future__ import annotations

import io
import json
import math
import os
import random
import string
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import requests
from flask import Flask, jsonify, redirect, render_template, request, send_file, session
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from urllib.parse import quote, unquote
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ModuleNotFoundError:  # openpyxl 미설치 환경
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore
    OPENPYXL_AVAILABLE = False

    def get_column_letter(_n: int) -> str:  # pragma: no cover
        # openpyxl이 없을 때는 _autosize_columns가 호출되지 않도록 가드함
        return "A"



# -------------------------
# Database bootstrap
# -------------------------

DATABASE_URL = (os.environ.get("DATABASE_URL") or "").strip()
if DATABASE_URL.startswith("postgres://"):
    # Render/Heroku style legacy scheme
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = None
SessionLocal = None

if DATABASE_URL:
    engine = create_engine(DATABASE_URL, pool_pre_ping=True)
    SessionLocal = sessionmaker(bind=engine)


def init_db() -> None:
    """
    DB schema initialization + minimal migration skeleton.

    How it works:
    - A single-row table schema_migrations stores the current schema version.
    - migrations is a list of (version, [sql...]) applied in order.
    - To evolve schema later WITHOUT dropping data:
        - Add a new (version+1, [...ALTER/CREATE...]) entry
        - Deploy; init_db will apply missing migrations.

    Dev mode (data not important):
    - Dropping tables and letting init_db recreate is simplest.
    """
    if not engine:
        return

    with engine.begin() as conn:
        # 0) schema_migrations
        conn.execute(text("""
        CREATE TABLE IF NOT EXISTS schema_migrations (
            id INTEGER PRIMARY KEY,
            version INTEGER NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW()
        );
        """))
        conn.execute(text("""
        INSERT INTO schema_migrations (id, version)
        VALUES (1, 0)
        ON CONFLICT (id) DO NOTHING;
        """))

        row = conn.execute(text("SELECT version FROM schema_migrations WHERE id = 1")).fetchone()
        current_version = int(row.version) if row and row.version is not None else 0

        def set_version(v: int) -> None:
            conn.execute(text("""
            UPDATE schema_migrations
            SET version = :v,
                updated_at = NOW()
            WHERE id = 1
            """), {"v": int(v)})

        migrations: List[Tuple[int, List[str]]] = []

        # -------------------
        # Migration v1: canonical schema for current app
        # -------------------
        migrations.append((1, [
            # teachers
            """
            CREATE TABLE IF NOT EXISTS teachers (
                id SERIAL PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,

            # classes
            """
            CREATE TABLE IF NOT EXISTS classes (
                id SERIAL PRIMARY KEY,
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                teacher_username TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,

            # students
            """
            CREATE TABLE IF NOT EXISTS students (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                student_no TEXT,
                name TEXT NOT NULL
            );
            """,

            # student_sessions (canonical: sid + placements JSONB)
            """
            CREATE TABLE IF NOT EXISTS student_sessions (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                sid TEXT NOT NULL,
                student_name TEXT NOT NULL,
                placements JSONB,
                placements_json TEXT,                -- legacy compatibility
                confidence INTEGER,
                priority INTEGER,
                submitted BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,
            "ALTER TABLE student_sessions ADD COLUMN IF NOT EXISTS sid TEXT;",
            "ALTER TABLE student_sessions ADD COLUMN IF NOT EXISTS student_name TEXT;",

            """
            CREATE UNIQUE INDEX IF NOT EXISTS uq_student_sessions
            ON student_sessions (class_code, sid, student_name);
            """,
            "ALTER TABLE student_sessions ADD COLUMN IF NOT EXISTS placements JSONB;",
            "ALTER TABLE student_sessions ADD COLUMN IF NOT EXISTS placements_json TEXT;",
            "ALTER TABLE student_sessions ADD COLUMN IF NOT EXISTS submitted BOOLEAN DEFAULT FALSE;",
            "ALTER TABLE student_sessions ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT NOW();",

            # teacher_placement_runs (canonical: session_id + placements JSONB + timing)
            """
            CREATE TABLE IF NOT EXISTS teacher_placement_runs (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                teacher_username TEXT NOT NULL,
                condition TEXT,
                tool_run_id INTEGER,
                placements JSONB,
                placements_json TEXT,                -- legacy compatibility
                submitted BOOLEAN DEFAULT FALSE,
                started_at TIMESTAMP DEFAULT NOW(),
                ended_at TIMESTAMP,
                duration_ms INTEGER,
                confidence_score INTEGER,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_teacher_runs_class_session
            ON teacher_placement_runs (class_code, session_id);
            """,
            "ALTER TABLE teacher_placement_runs ADD COLUMN IF NOT EXISTS condition TEXT;",
            "ALTER TABLE teacher_placement_runs ADD COLUMN IF NOT EXISTS tool_run_id INTEGER;",
            "ALTER TABLE teacher_placement_runs ADD COLUMN IF NOT EXISTS placements JSONB;",
            "ALTER TABLE teacher_placement_runs ADD COLUMN IF NOT EXISTS placements_json TEXT;",
            "ALTER TABLE teacher_placement_runs ADD COLUMN IF NOT EXISTS submitted BOOLEAN DEFAULT FALSE;",
            "ALTER TABLE teacher_placement_runs ADD COLUMN IF NOT EXISTS started_at TIMESTAMP DEFAULT NOW();",
            "ALTER TABLE teacher_placement_runs ADD COLUMN IF NOT EXISTS ended_at TIMESTAMP;",
            "ALTER TABLE teacher_placement_runs ADD COLUMN IF NOT EXISTS duration_ms INTEGER;",
            "ALTER TABLE teacher_placement_runs ADD COLUMN IF NOT EXISTS confidence_score INTEGER;",
            "ALTER TABLE teacher_placement_runs ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT NOW();",

            # teacher_decisions (canonical: run_id rows)
            """
            CREATE TABLE IF NOT EXISTS teacher_decisions (
                id SERIAL PRIMARY KEY,
                run_id INTEGER NOT NULL,
                target_student_name TEXT NOT NULL,
                priority_rank INTEGER NOT NULL,
                decision_confidence INTEGER,
                reason_tags JSONB,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_teacher_decisions_run
            ON teacher_decisions (run_id);
            """,
            "ALTER TABLE teacher_decisions ADD COLUMN IF NOT EXISTS decision_confidence INTEGER;",
            "ALTER TABLE teacher_decisions ADD COLUMN IF NOT EXISTS reason_tags JSONB;",
            "ALTER TABLE teacher_decisions ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT NOW();",

            # analysis_cache
            """
            CREATE TABLE IF NOT EXISTS analysis_cache (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                cache_key TEXT NOT NULL,
                payload JSONB,
                updated_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(class_code, session_id, cache_key)
            );
            """,
            "ALTER TABLE analysis_cache ADD COLUMN IF NOT EXISTS payload JSONB;",
            "ALTER TABLE analysis_cache ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT NOW();",
        ]))

        # -------------------
        # Migration v2: roster(gender/pin/active) + finalize + artifacts
        # -------------------
        migrations.append((2, [
            # --- 학생 관리(전입/전출/성별/개인코드) ---
            "ALTER TABLE students ADD COLUMN IF NOT EXISTS gender TEXT;",
            "ALTER TABLE students ADD COLUMN IF NOT EXISTS pin_code TEXT;",
            "ALTER TABLE students ADD COLUMN IF NOT EXISTS active BOOLEAN DEFAULT TRUE;",
            "ALTER TABLE students ADD COLUMN IF NOT EXISTS joined_at TIMESTAMP DEFAULT NOW();",
            "ALTER TABLE students ADD COLUMN IF NOT EXISTS left_at TIMESTAMP;",

            # 학급 내 pin 중복 방지(학급 단위로만 유일하면 됨)
            """
            CREATE UNIQUE INDEX IF NOT EXISTS uq_students_class_pin
            ON students (class_code, pin_code);
            """,

            # --- 회차 마무리 상태(회차당 1행) ---
            """
            CREATE TABLE IF NOT EXISTS session_finalizations (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                teacher_username TEXT NOT NULL,

                preview_seen BOOLEAN DEFAULT FALSE,
                preview_seen_at TIMESTAMP,

                exclusions_resolved BOOLEAN DEFAULT FALSE,
                exclusions_resolved_at TIMESTAMP,

                survey_submitted BOOLEAN DEFAULT FALSE,
                survey_submitted_at TIMESTAMP,

                finalized BOOLEAN DEFAULT FALSE,
                finalized_at TIMESTAMP,

                updated_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(class_code, session_id)
            );
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_session_finalizations_class_session
            ON session_finalizations (class_code, session_id);
            """,

            # --- 참여 제외 기록(학생 단위) ---
            """
            CREATE TABLE IF NOT EXISTS session_exclusions (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                student_name TEXT NOT NULL,
                excluded BOOLEAN DEFAULT TRUE,
                reason TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(class_code, session_id, student_name)
            );
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_session_exclusions_class_session
            ON session_exclusions (class_code, session_id);
            """,

            # --- 교사 설문(회차당 1행) ---
            """
            CREATE TABLE IF NOT EXISTS teacher_surveys (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                teacher_username TEXT NOT NULL,

                q1_help INTEGER,
                q2_new TEXT,
                q2_detail TEXT,
                q3_use INTEGER,
                q4_cmp TEXT,
                q4_detail TEXT,
                q5_conf TEXT,
                q6_next TEXT,
                q7_feedback TEXT,

                submitted_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(class_code, session_id)
            );
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_teacher_surveys_class_session
            ON teacher_surveys (class_code, session_id);
            """,

            # --- 회차 결과 스냅샷(논문 재현성/추적용) ---
            """
            CREATE TABLE IF NOT EXISTS session_artifacts (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                artifact_key TEXT NOT NULL,
                artifact_version TEXT NOT NULL,
                params JSONB,
                payload JSONB NOT NULL,
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(class_code, session_id, artifact_key)
            );
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_session_artifacts_class_session
            ON session_artifacts (class_code, session_id);
            """,
        ]))


        migrations.sort(key=lambda x: int(x[0]))
        for target_version, stmts in migrations:
            if current_version >= int(target_version):
                continue
            for stmt in stmts:
                if not stmt or not str(stmt).strip():
                    continue
                conn.execute(text(stmt))
            set_version(int(target_version))
            current_version = int(target_version)


# Initialize DB tables on startup (best-effort; app still boots for debug)
try:
    init_db()
except Exception as e:
    print("init_db failed:", e)

# -------------------------
# Flask bootstrap
# -------------------------

app = Flask(__name__)
app.config["PROPAGATE_EXCEPTIONS"] = True
app.secret_key = os.environ.get("SECRET_KEY", "dev-only-change-me")

app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"


@app.before_request
def make_session_permanent() -> None:
    session.permanent = True

# -------------------------
# Research admin (owner-only)
# -------------------------
ADMIN_USERS = {
    u.strip()
    for u in (os.environ.get("ADMIN_USERS", "") or "").split(",")
    if u.strip()
}

def require_admin():
    """
    Owner-only research/admin pages.
    - Must be logged in as teacher.
    - Username must be in ADMIN_USERS (env var, comma-separated).
    """
    if "teacher" not in session:
        return redirect("/teacher/login")

    if ADMIN_USERS and session.get("teacher") not in ADMIN_USERS:
        # 권한 없음
        return "forbidden", 403

    # ADMIN_USERS가 비어있으면(환경변수 미설정) 안전하게 막고 싶다면 아래 주석 해제:
    # if not ADMIN_USERS:
    #     return "forbidden (ADMIN_USERS not set)", 403

    return None

# -------------------------
# Research admin: XLSX helpers + overview fetch
# -------------------------

def _xlsx_response(wb, filename: str):
    """
    Safe XLSX response helper.
    - openpyxl 미설치/오작동 시: 라우트가 500으로 깔끔하게 실패하고, 앱 전체 부팅을 깨지 않도록 방어
    """
    if not OPENPYXL_AVAILABLE or Workbook is None:
        return "openpyxl not installed on server", 500

    if wb is None:
        return "workbook is None", 500

    bio = io.BytesIO()
    try:
        wb.save(bio)
    except Exception as e:
        return f"failed to generate xlsx: {e}", 500

    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _autosize_columns(ws):
    # 단순 자동 폭 (완벽하진 않지만 연구용엔 충분)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
            except Exception:
                v = ""
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def db_fetch_class_overview() -> List[Dict[str, Any]]:
    """
    연구/관리 페이지에서 '지금 바로 연결 가능한' 요약 데이터:
    - classes 목록
    - 학생 수
    - student_sessions: sid별 제출/전체
    - teacher_placement_runs: session_id별 제출/전체
    """
    if not engine:
        return []

    with engine.connect() as conn:
        classes = conn.execute(text("""
            SELECT code, name, teacher_username, created_at
            FROM classes
            ORDER BY id DESC
        """)).fetchall()

        out: List[Dict[str, Any]] = []
        for c in classes:
            student_cnt = conn.execute(text("""
                SELECT COUNT(*) AS n FROM students WHERE class_code = :code
            """), {"code": c.code}).fetchone().n

            ss_rows = conn.execute(text("""
                SELECT sid,
                       COUNT(*) AS total,
                       SUM(CASE WHEN submitted THEN 1 ELSE 0 END) AS submitted
                FROM student_sessions
                WHERE class_code = :code
                GROUP BY sid
                ORDER BY sid::int
            """), {"code": c.code}).fetchall()

            tr_rows = conn.execute(text("""
                SELECT session_id,
                       COUNT(*) AS total,
                       SUM(CASE WHEN submitted THEN 1 ELSE 0 END) AS submitted
                FROM teacher_placement_runs
                WHERE class_code = :code
                GROUP BY session_id
                ORDER BY session_id::int
            """), {"code": c.code}).fetchall()

            out.append({
                "code": c.code,
                "name": c.name,
                "teacher_username": c.teacher_username,
                "created_at": c.created_at,
                "student_count": int(student_cnt or 0),
                "student_sessions": [
                    {"sid": r.sid, "total": int(r.total or 0), "submitted": int(r.submitted or 0)}
                    for r in ss_rows
                ],
                "teacher_runs": [
                    {"session_id": r.session_id, "total": int(r.total or 0), "submitted": int(r.submitted or 0)}
                    for r in tr_rows
                ],
            })
        return out


SITE_TITLE = "내가 바라본 우리 반"

# JSON fallback file (only used if DB not configured)
DATA_FILE = os.environ.get("DATA_FILE", "data.json")

# Google Sheets integration
GOOGLE_WEBAPP_URL = os.environ.get(
    "GOOGLE_WEBAPP_URL",
    "https://script.google.com/macros/s/AKfycbwyjKC2JearJnySkxdG0oahMkMJ5V6uBqY5EYRGVVRa8KWZvRzHcskeVNY5hnlyiSw/exec",
)
GOOGLE_SECRET = (os.environ.get("GOOGLE_SECRET") or "").strip()

DEBUG_MODE = os.environ.get("DEBUG_MODE") == "1"


# -------------------------
# Utilities: Google Sheets
# -------------------------

def post_to_sheet(payload: Dict[str, Any]) -> Dict[str, Any]:
    payload = dict(payload)
    payload["secret"] = GOOGLE_SECRET

    try:
        r = requests.post(GOOGLE_WEBAPP_URL, json=payload, timeout=10)
    except Exception as e:
        return {"status": "error", "message": f"request failed: {e}"}

    if r.status_code != 200:
        return {"status": "error", "message": f"http {r.status_code}", "text": r.text[:300]}

    try:
        return r.json()
    except Exception:
        return {"status": "error", "message": "invalid json response", "text": r.text[:300]}


# -------------------------
# Utilities: JSON file fallback
# -------------------------

def ensure_class_schema(cls: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not cls:
        return cls
    cls.setdefault("sessions", {})
    for i in range(1, 6):
        cls["sessions"].setdefault(str(i), {"label": f"{i}차", "active": i == 1})

    for name, sdata in cls.get("students_data", {}).items():
        sdata.setdefault("sessions", {})
        for sid in cls["sessions"]:
            sdata["sessions"].setdefault(sid, {"placements": {}, "submitted": False})
    return cls


def load_data() -> Dict[str, Any]:
    if not os.path.exists(DATA_FILE):
        return {"classes": {}}

    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            d = json.load(f)
    except Exception:
        return {"classes": {}}

    d.setdefault("classes", {})
    for code, cls in list(d.get("classes", {}).items()):
        d["classes"][code] = ensure_class_schema(cls)
    return d


def save_data(data: Dict[str, Any]) -> None:
    parent_dir = os.path.dirname(DATA_FILE) or "."
    os.makedirs(parent_dir, exist_ok=True)

    tmp_path = DATA_FILE + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, DATA_FILE)


def save_data_safely(d: Dict[str, Any]) -> None:
    for code, cls in list(d.get("classes", {}).items()):
        d["classes"][code] = ensure_class_schema(cls)
    save_data(d)


def make_code() -> str:
    return "".join(random.choices(string.ascii_uppercase + string.digits, k=6))


# -------------------------
# DB helpers (canonical schema)
# -------------------------

def _json_load_maybe(val: Any) -> Dict[str, Any]:
    if isinstance(val, dict):
        return val
    if isinstance(val, str):
        try:
            obj = json.loads(val)
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}
    return {}


def db_list_classes_for_teacher(teacher_username: str) -> Dict[str, Dict[str, str]]:
    if not engine:
        raise RuntimeError("DB engine not initialized")
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT code, name, teacher_username
            FROM classes
            WHERE teacher_username = :t
            ORDER BY id DESC
        """), {"t": teacher_username}).fetchall()

    out: Dict[str, Dict[str, str]] = {}
    for r in rows:
        out[r.code] = {"name": r.name, "teacher": r.teacher_username}
    return out


def db_create_class(teacher_username: str, class_code: str, class_name: str, students: List[Dict[str, str]]) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    import random

    def make_pin(existing: set) -> str:
        # 6자리 숫자: 학급 내부에서만 유일하면 됨
        # (000000도 가능하지만 헷갈리므로 100000~999999 권장)
        for _ in range(2000):
            pin = str(random.randint(100000, 999999))
            if pin not in existing:
                existing.add(pin)
                return pin
        # 최악의 경우(거의 없음): 순차 발급
        pin = "100000"
        while pin in existing:
            pin = str(int(pin) + 1)
        existing.add(pin)
        return pin

    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO classes (code, name, teacher_username)
            VALUES (:code, :name, :t)
        """), {"code": class_code, "name": class_name, "t": teacher_username})

        existing_pins: set = set()

        for s in students:
            nm = (s.get("name") or "").strip()
            if not nm:
                continue

            pin = make_pin(existing_pins)
            gender = (s.get("gender") or "").upper()
            if gender not in ["M", "F"]:
                gender = ""

            conn.execute(text("""
                INSERT INTO students (class_code, student_no, name, gender, pin_code, active, joined_at)
                VALUES (:code, :no, :name, :gender, :pin, TRUE, NOW())
            """), {
                "code": class_code,
                "no": str(s.get("no", "") or ""),
                "name": nm,
                "gender": gender,
                "pin": pin,
            })



def db_delete_class_for_teacher(class_code: str, teacher_username: str) -> bool:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    class_code = (class_code or "").upper().strip()

    with engine.begin() as conn:
        row = conn.execute(text("""
            SELECT 1 FROM classes
            WHERE code = :code AND teacher_username = :t
            LIMIT 1
        """), {"code": class_code, "t": teacher_username}).fetchone()
        if not row:
            return False

        conn.execute(text("DELETE FROM teacher_decisions WHERE run_id IN (SELECT id FROM teacher_placement_runs WHERE class_code = :code)"), {"code": class_code})
        conn.execute(text("DELETE FROM teacher_placement_runs WHERE class_code = :code"), {"code": class_code})
        conn.execute(text("DELETE FROM student_sessions WHERE class_code = :code"), {"code": class_code})
        conn.execute(text("DELETE FROM students WHERE class_code = :code"), {"code": class_code})
        conn.execute(text("DELETE FROM analysis_cache WHERE class_code = :code"), {"code": class_code})
        conn.execute(text("DELETE FROM classes WHERE code = :code AND teacher_username = :t"), {"code": class_code, "t": teacher_username})

    return True


def db_get_class_name(class_code: str) -> Optional[str]:
    if not engine:
        return None
    with engine.connect() as conn:
        row = conn.execute(text("SELECT name FROM classes WHERE code = :code LIMIT 1"), {"code": class_code}).fetchone()
    return row.name if row else None


def db_get_class_for_teacher(class_code: str, teacher_username: str) -> Optional[Dict[str, Any]]:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT code, name, teacher_username
            FROM classes
            WHERE code = :code
            LIMIT 1
        """), {"code": class_code}).fetchone()

    if not row:
        return None
    if row.teacher_username != teacher_username:
        return {"_forbidden": True}

    return {"code": row.code, "name": row.name, "teacher": row.teacher_username, "sessions": {}}


def db_get_students_in_class(class_code: str) -> List[Dict[str, str]]:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT student_no, name
            FROM students
            WHERE class_code = :code
            ORDER BY id ASC
        """), {"code": class_code}).fetchall()

    out: List[Dict[str, str]] = []
    for r in rows:
        out.append({"no": (r.student_no or ""), "name": r.name})
    return out


def db_get_submitted_map(class_code: str, sid: str) -> Dict[str, bool]:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT student_name, submitted
            FROM student_sessions
            WHERE class_code = :code AND sid = :sid
        """), {"code": class_code, "sid": sid}).fetchall()

    m: Dict[str, bool] = {}
    for r in rows:
        m[r.student_name] = bool(r.submitted)
    return m



def db_get_session_state_v2(class_code: str, sid: str, teacher_username: str) -> Dict[str, Any]:
    """Return v2 finalize state + exclusions + survey flags for a given class/session."""
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        # Ensure anchor row exists
        conn.execute(text("""
            INSERT INTO session_finalizations (class_code, session_id, teacher_username)
            VALUES (:code, :sid, :t)
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET teacher_username = EXCLUDED.teacher_username,
                          updated_at = NOW()
        """), {"code": class_code, "sid": sid, "t": teacher_username})

        fin = conn.execute(text("""
            SELECT preview_seen, exclusions_resolved, survey_submitted, finalized,
                   preview_seen_at, exclusions_resolved_at, survey_submitted_at, finalized_at
            FROM session_finalizations
            WHERE class_code = :code AND session_id = :sid
        """), {"code": class_code, "sid": sid}).fetchone()

        exc_rows = conn.execute(text("""
            SELECT student_name, excluded, reason
            FROM session_exclusions
            WHERE class_code = :code AND session_id = :sid
            ORDER BY student_name
        """), {"code": class_code, "sid": sid}).fetchall()

        survey_row = conn.execute(text("""
            SELECT 1 AS ok
            FROM teacher_surveys
            WHERE class_code = :code AND session_id = :sid
        """), {"code": class_code, "sid": sid}).fetchone()

    exclusions = []
    for r in exc_rows:
        exclusions.append({
            "student_name": r.student_name,
            "excluded": bool(r.excluded),
            "reason": (r.reason or "")
        })

    state = {
        "class_code": class_code,
        "session_id": sid,
        "preview_seen": bool(getattr(fin, "preview_seen", False)),
        "exclusions_resolved": bool(getattr(fin, "exclusions_resolved", False)),
        "survey_submitted": bool(getattr(fin, "survey_submitted", False)) or bool(survey_row),
        "finalized": bool(getattr(fin, "finalized", False)),
        "exclusions": exclusions,
    }
    return state


def db_set_preview_seen_v2(class_code: str, sid: str, teacher_username: str) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO session_finalizations (class_code, session_id, teacher_username, preview_seen, preview_seen_at)
            VALUES (:code, :sid, :t, TRUE, NOW())
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET preview_seen = TRUE,
                          preview_seen_at = NOW(),
                          teacher_username = EXCLUDED.teacher_username,
                          updated_at = NOW()
        """), {"code": class_code, "sid": sid, "t": teacher_username})


def db_set_exclusions_v2(class_code: str, sid: str, teacher_username: str, mode: str, items: List[Dict[str, str]]) -> None:
    """mode:
      - 'no_exclude': mark exclusions_resolved true, keep exclusions table empty
      - 'exclude': upsert exclusions with reasons, mark exclusions_resolved true
    """
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        # Clear previous exclusions for this session (simple and safe)
        conn.execute(text("""
            DELETE FROM session_exclusions
            WHERE class_code = :code AND session_id = :sid
        """), {"code": class_code, "sid": sid})

        if mode == "exclude":
            for it in items:
                nm = (it.get("student_name") or "").strip()
                rs = (it.get("reason") or "").strip()
                if not nm:
                    continue
                conn.execute(text("""
                    INSERT INTO session_exclusions (class_code, session_id, student_name, excluded, reason)
                    VALUES (:code, :sid, :nm, TRUE, :rs)
                    ON CONFLICT (class_code, session_id, student_name)
                    DO UPDATE SET excluded = TRUE, reason = EXCLUDED.reason
                """), {"code": class_code, "sid": sid, "nm": nm, "rs": rs})

        conn.execute(text("""
            INSERT INTO session_finalizations (class_code, session_id, teacher_username, exclusions_resolved, exclusions_resolved_at)
            VALUES (:code, :sid, :t, TRUE, NOW())
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET exclusions_resolved = TRUE,
                          exclusions_resolved_at = NOW(),
                          teacher_username = EXCLUDED.teacher_username,
                          updated_at = NOW()
        """), {"code": class_code, "sid": sid, "t": teacher_username})


def db_upsert_survey_v2(class_code: str, sid: str, teacher_username: str, payload: Dict[str, Any]) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    def _to_int(v: Any) -> Optional[int]:
        try:
            if v is None or v == "":
                return None
            return int(v)
        except Exception:
            return None

    row = {
        "code": class_code,
        "sid": sid,
        "t": teacher_username,
        "q1_help": _to_int(payload.get("q1_help")),
        "q2_new": (payload.get("q2_new") or None),
        "q2_detail": (payload.get("q2_detail") or None),
        "q3_use": _to_int(payload.get("q3_use")),
        "q4_cmp": (payload.get("q4_cmp") or None),
        "q4_detail": (payload.get("q4_detail") or None),
        "q5_conf": (payload.get("q5_conf") or None),
        "q6_next": (payload.get("q6_next") or None),
        "q7_feedback": (payload.get("q7_feedback") or None),
    }

    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO teacher_surveys
            (class_code, session_id, teacher_username,
             q1_help, q2_new, q2_detail, q3_use, q4_cmp, q4_detail, q5_conf, q6_next, q7_feedback,
             submitted_at)
            VALUES
            (:code, :sid, :t,
             :q1_help, :q2_new, :q2_detail, :q3_use, :q4_cmp, :q4_detail, :q5_conf, :q6_next, :q7_feedback,
             NOW())
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET
              teacher_username = EXCLUDED.teacher_username,
              q1_help = EXCLUDED.q1_help,
              q2_new = EXCLUDED.q2_new,
              q2_detail = EXCLUDED.q2_detail,
              q3_use = EXCLUDED.q3_use,
              q4_cmp = EXCLUDED.q4_cmp,
              q4_detail = EXCLUDED.q4_detail,
              q5_conf = EXCLUDED.q5_conf,
              q6_next = EXCLUDED.q6_next,
              q7_feedback = EXCLUDED.q7_feedback,
              submitted_at = NOW()
        """), row)

        conn.execute(text("""
            INSERT INTO session_finalizations (class_code, session_id, teacher_username, survey_submitted, survey_submitted_at)
            VALUES (:code, :sid, :t, TRUE, NOW())
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET survey_submitted = TRUE,
                          survey_submitted_at = NOW(),
                          teacher_username = EXCLUDED.teacher_username,
                          updated_at = NOW()
        """), {"code": class_code, "sid": sid, "t": teacher_username})


def db_finalize_session_v2(class_code: str, sid: str, teacher_username: str) -> Dict[str, Any]:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        fin = conn.execute(text("""
            SELECT preview_seen, exclusions_resolved, survey_submitted, finalized
            FROM session_finalizations
            WHERE class_code = :code AND session_id = :sid
        """), {"code": class_code, "sid": sid}).fetchone()

        if not fin:
            raise ValueError("Finalize state not found")

        preview_seen = bool(fin.preview_seen)
        exclusions_resolved = bool(fin.exclusions_resolved)
        survey_submitted = bool(fin.survey_submitted) or bool(conn.execute(text("""
            SELECT 1 FROM teacher_surveys WHERE class_code=:code AND session_id=:sid
        """), {"code": class_code, "sid": sid}).fetchone())

        if not (preview_seen and exclusions_resolved and survey_submitted):
            return {
                "ok": False,
                "error": "NOT_READY",
                "preview_seen": preview_seen,
                "exclusions_resolved": exclusions_resolved,
                "survey_submitted": survey_submitted,
            }

        conn.execute(text("""
            UPDATE session_finalizations
            SET finalized = TRUE,
                finalized_at = NOW(),
                teacher_username = :t,
                updated_at = NOW()
            WHERE class_code = :code AND session_id = :sid
        """), {"code": class_code, "sid": sid, "t": teacher_username})

        return {"ok": True, "finalized": True}


def db_get_student_session(class_code: str, student_name: str, sid: str) -> Optional[Dict[str, Any]]:
    if not engine:
        return None

    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT placements, placements_json, submitted
            FROM student_sessions
            WHERE class_code = :code
              AND student_name = :name
              AND sid = :sid
            LIMIT 1
        """), {"code": class_code, "name": student_name, "sid": sid}).fetchone()

    if not row:
        return None

    placements_obj: Dict[str, Any] = {}
    if row.placements is not None:
        placements_obj = row.placements if isinstance(row.placements, dict) else _json_load_maybe(row.placements)
    elif row.placements_json:
        placements_obj = _json_load_maybe(row.placements_json)

    return {"placements": placements_obj, "submitted": bool(row.submitted)}


def db_list_submitted_student_sessions(class_code: str, sid: str) -> List[Dict[str, Any]]:
    if not engine:
        return []

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT student_name, placements, placements_json
            FROM student_sessions
            WHERE class_code = :code
              AND sid = :sid
              AND submitted = TRUE
            ORDER BY id ASC
        """), {"code": class_code, "sid": sid}).fetchall()

    out: List[Dict[str, Any]] = []
    for r in rows:
        placements_obj: Dict[str, Any] = {}
        if r.placements is not None:
            placements_obj = r.placements if isinstance(r.placements, dict) else _json_load_maybe(r.placements)
        elif r.placements_json:
            placements_obj = _json_load_maybe(r.placements_json)
        out.append({"student_name": r.student_name, "placements": placements_obj})
    return out

def db_upsert_student_session(class_code: str, student_name: str, sid: str, placements: Dict[str, Any], submitted: bool) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    placements_str = json.dumps(placements, ensure_ascii=False)

    with engine.begin() as conn:
        row = conn.execute(text("""
            SELECT id
            FROM student_sessions
            WHERE class_code = :code AND student_name = :name AND sid = :sid
            LIMIT 1
        """), {"code": class_code, "name": student_name, "sid": sid}).fetchone()

        if row:
            conn.execute(text("""
                UPDATE student_sessions
                SET placements = CAST(:placements AS jsonb),
                    placements_json = :placements_json,
                    submitted = :submitted
                WHERE id = :id
            """), {"placements": placements_str, "placements_json": placements_str, "submitted": submitted, "id": row.id})
        else:
            conn.execute(text("""
                INSERT INTO student_sessions (class_code, sid, student_name, placements, placements_json, submitted)
                VALUES (:code, :sid, :name, CAST(:placements AS jsonb), :placements_json, :submitted)
            """), {"code": class_code, "sid": sid, "name": student_name, "placements": placements_str, "placements_json": placements_str, "submitted": submitted})


def db_create_teacher_run(class_code: str, teacher_username: str, sid: str, condition: str, tool_run_id: Optional[int] = None) -> int:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        row = conn.execute(text("""
            INSERT INTO teacher_placement_runs
            (class_code, teacher_username, session_id, condition, tool_run_id, placements, placements_json, submitted, started_at)
            VALUES (:code, :t, :sid, :cond, :tool_run_id, CAST(:placements AS jsonb), :placements_json, FALSE, NOW())
            RETURNING id
        """), {
            "code": class_code,
            "t": teacher_username,
            "sid": sid,
            "cond": condition,
            "tool_run_id": tool_run_id,
            "placements": json.dumps({}, ensure_ascii=False),
            "placements_json": json.dumps({}, ensure_ascii=False),
        }).fetchone()

    return int(row[0])




def db_get_teacher_run(run_id: int) -> Optional[Dict[str, Any]]:
    if not engine:
        return None

    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT id, class_code, teacher_username, session_id, condition, tool_run_id,
                   placements, placements_json, submitted, started_at, ended_at, duration_ms, confidence_score
            FROM teacher_placement_runs
            WHERE id = :id
            LIMIT 1
        """), {"id": run_id}).fetchone()

    if not row:
        return None

    placements_obj: Dict[str, Any] = {}
    if row.placements is not None:
        placements_obj = row.placements if isinstance(row.placements, dict) else _json_load_maybe(row.placements)
    elif row.placements_json:
        placements_obj = _json_load_maybe(row.placements_json)

    return {
        "id": row.id,
        "class_code": row.class_code,
        "teacher_username": row.teacher_username,
        "session_id": row.session_id,
        "condition": row.condition,
        "tool_run_id": row.tool_run_id,
        "placements": placements_obj,
        "submitted": bool(row.submitted),
        "started_at": row.started_at,
        "ended_at": row.ended_at,
        "duration_ms": row.duration_ms,
        "confidence_score": row.confidence_score,
    }


def db_update_teacher_run_placements(run_id: int, placements: Dict[str, Any]) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    placements_str = json.dumps(placements, ensure_ascii=False)

    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE teacher_placement_runs
            SET placements = CAST(:placements AS jsonb),
                placements_json = :placements_json
            WHERE id = :id
        """), {"placements": placements_str, "placements_json": placements_str, "id": run_id})



def db_complete_teacher_run(run_id: int, duration_ms: int, confidence_score: int) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE teacher_placement_runs
            SET ended_at = NOW(),
                duration_ms = :duration_ms,
                confidence_score = :confidence_score,
                submitted = TRUE
            WHERE id = :id
        """), {"duration_ms": int(duration_ms), "confidence_score": int(confidence_score), "id": run_id})


def db_replace_teacher_decisions(run_id: int, decisions: List[Dict[str, Any]]) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM teacher_decisions WHERE run_id = :run_id"), {"run_id": run_id})
        for d in decisions:
            conn.execute(text("""
                INSERT INTO teacher_decisions
                (run_id, target_student_name, priority_rank, decision_confidence, reason_tags)
                VALUES (:run_id, :name, :rank, :conf, CAST(:tags AS jsonb))
            """), {
                "run_id": run_id,
                "name": (d.get("name") or "").strip(),
                "rank": int(d.get("rank") or 0),
                "conf": int(d.get("confidence") or 0) if d.get("confidence") is not None else None,
                "tags": json.dumps(d.get("tags") or [], ensure_ascii=False),
            })



# -------------------------
# Cache helpers (analysis_cache)
# -------------------------

def cache_get(class_code: str, sid: str, key: str) -> Optional[Dict[str, Any]]:
    if not engine:
        return None
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT payload
            FROM analysis_cache
            WHERE class_code = :c AND session_id = :s AND cache_key = :k
            LIMIT 1
        """), {"c": class_code, "s": sid, "k": key}).fetchone()
    if not row:
        return None
    if isinstance(row.payload, dict):
        return row.payload
    return _json_load_maybe(row.payload)


def cache_set(class_code: str, sid: str, key: str, payload: Dict[str, Any]) -> None:
    if not engine:
        return
    payload_str = json.dumps(payload, ensure_ascii=False)
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO analysis_cache (class_code, session_id, cache_key, payload, updated_at)
            VALUES (:c, :s, :k, CAST(:p AS jsonb), NOW())
            ON CONFLICT (class_code, session_id, cache_key)
            DO UPDATE SET payload = EXCLUDED.payload, updated_at = NOW()
        """), {"c": class_code, "s": sid, "k": key, "p": payload_str})



# -------------------------
# Context: current class (topbar)
# -------------------------

def get_current_class() -> Optional[Dict[str, str]]:
    code = None
    if "teacher" in session and session.get("selected_class"):
        code = session.get("selected_class")
    elif session.get("code"):
        code = session.get("code")

    if not code:
        return None

    if engine:
        name = db_get_class_name(code)
        if not name:
            return None
        return {"name": name, "code": code}

    d = load_data()
    cls = d.get("classes", {}).get(code)
    if not cls:
        return None
    return {"name": cls.get("name", ""), "code": code}


@app.context_processor
def inject_globals() -> Dict[str, Any]:
    return {"current_class": get_current_class()}


# -------------------------
# Debug routes
# -------------------------

@app.route("/debug/db")
def debug_db():
    if not engine:
        return "DATABASE_URL not set"
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return "DB connection OK"
    except Exception as e:
        return f"DB connection failed: {e}", 500


if DEBUG_MODE:

    @app.route("/debug/schema_version")
    def debug_schema_version():
        if not engine:
            return jsonify({"error": "DATABASE_URL not set"}), 500
        try:
            with engine.connect() as conn:
                row = conn.execute(
                    text("SELECT version, updated_at FROM schema_migrations WHERE id = 1")
                ).fetchone()
            if not row:
                return jsonify({"version": None, "updated_at": None})
            return jsonify({
                "version": int(row.version),
                "updated_at": row.updated_at.isoformat() if row.updated_at else None,
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 500


    @app.route("/debug/versions")
    def debug_versions():
        import sys

        out = {
            "python_version": sys.version,
            "python_version_info": list(sys.version_info),
            "openpyxl_available": bool(OPENPYXL_AVAILABLE and (Workbook is not None)),
        }

        # openpyxl 버전도 확인 (설치돼 있을 때만)
        if OPENPYXL_AVAILABLE:
            try:
                import openpyxl  # type: ignore
                out["openpyxl_version"] = getattr(openpyxl, "__version__", None)
            except Exception as e:
                out["openpyxl_version_error"] = str(e)

        return jsonify(out)



# ---------- 헬스 체크 (콜드 스타트 방지용) ----------
@app.route("/health")
def health():
    # 기본 응답은 가볍게 유지하되, 내부적으로는 핵심 의존성 상태를 함께 확인 가능
    status = {
        "ok": True,
        "openpyxl_available": bool(OPENPYXL_AVAILABLE and (Workbook is not None)),
    }

    # DB는 설정돼 있으면 가볍게 ping (실패해도 앱 전체는 살리고 상태만 표시)
    if engine:
        try:
            with engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            status["db_ok"] = True
        except Exception as e:
            status["db_ok"] = False
            status["db_error"] = str(e)
    else:
        status["db_ok"] = False
        status["db_error"] = "DATABASE_URL not set"

    # openpyxl이 “import 됐고 Workbook이 살아있는지”가 핵심
    http_code = 200 if status["openpyxl_available"] else 500
    return jsonify(status), http_code




# -------------------------
# Home
# -------------------------

@app.route("/")
def home():
    # 로그인 유지 중이면 home -> dashboard 흐름으로 고정
    if session.get("teacher"):
        return redirect("/teacher/dashboard")

    return render_template("home.html", site_title=SITE_TITLE)



# -------------------------
# Teacher auth
# -------------------------

@app.route("/teacher/signup", methods=["GET", "POST"])
def teacher_signup():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        pw = request.form.get("password", "")
        pw2 = request.form.get("password2", "")

        if not username or not pw:
            return render_template("teacher_signup.html", error="아이디/비밀번호를 입력해 주세요.")
        if pw != pw2:
            return render_template("teacher_signup.html", error="비밀번호가 서로 다릅니다.")

        # =========================================================
        # RESEARCH_ONLY_BEGIN
        # 논문 작성용 추가 수집(추후 삭제 예정)
        # =========================================================
        teaching_years = request.form.get("teaching_years", "").strip()
        current_grade = request.form.get("current_grade", "").strip()
        research_name = request.form.get("research_name", "").strip()
        research_school = request.form.get("research_school", "").strip()

        if not teaching_years:
            return render_template("teacher_signup.html", error="교직경력을 선택해 주세요.")
        if not current_grade:
            return render_template("teacher_signup.html", error="담당 학년을 선택해 주세요.")
        if not research_name:
            return render_template("teacher_signup.html", error="이름을 입력해 주세요.")
        if not research_school:
            return render_template("teacher_signup.html", error="학교를 입력해 주세요.")
        # =========================================================
        # RESEARCH_ONLY_END
        # =========================================================

        pw_hash = generate_password_hash(pw)

        try:
            resp = post_to_sheet({
                "action": "teacher_signup",
                "username": username,
                "pw_hash": pw_hash,

                # =========================================================
                # RESEARCH_ONLY_BEGIN
                # 논문 작성용 추가 수집(추후 삭제 예정)
                # =========================================================
                "teaching_years": teaching_years,      # "1-5" / "6-10" / "11-15" / "16+"
                "current_grade": current_grade,        # "4" / "5" / "6"
                "research_name": research_name,        # 필수
                "research_school": research_school,    # 필수
                # =========================================================
                # RESEARCH_ONLY_END
                # =========================================================
            })
        except Exception as e:
            return render_template("teacher_signup.html", error=f"서버 통신 오류: {e}")

        status = resp.get("status")
        if status == "ok":
            return redirect("/teacher/login")
        if status == "exists":
            return render_template("teacher_signup.html", error="이미 존재하는 아이디입니다.")
        if status == "blocked":
            return render_template("teacher_signup.html", error="blocked: GOOGLE_SECRET(비밀키) 불일치 또는 누락")
        return render_template("teacher_signup.html", error=f"회원가입 실패: {resp}")

    return render_template("teacher_signup.html")


@app.route("/teacher/login", methods=["GET", "POST"])
def teacher_login():
    # 로그인 유지 중이면 login 화면을 거치지 않고 home으로
    # (home이 다시 dashboard로 보내므로 경로가 home -> dashboard로 고정됨)
    if request.method == "GET" and session.get("teacher"):
        return redirect("/")

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        pw = request.form.get("password", "")

        try:
            resp = post_to_sheet({"action": "teacher_get", "username": username})
        except Exception as e:
            return render_template("teacher_login.html", error=f"서버 통신 오류: {e}")

        try:
            if resp.get("status") != "ok":
                return render_template("teacher_login.html", error=f"로그인 실패: {resp}")

            pw_hash = resp.get("pw_hash") or ""
            if check_password_hash(pw_hash, pw):
                session.clear()
                session["teacher"] = username

                # 바로 dashboard로 보내지 않고 home으로 보냄
                # (주소 흐름: home -> dashboard)
                return redirect("/")

            return render_template("teacher_login.html", error="로그인 실패: 비밀번호 불일치")
        except Exception as e:
            return render_template("teacher_login.html", error=f"로그인 처리 중 오류: {e} / resp={resp}")

    return render_template("teacher_login.html")




@app.route("/teacher/logout")
def teacher_logout():
    session.clear()
    return redirect("/")

# -------------------------
# Research admin pages (owner-only)
# -------------------------

@app.route("/research")
def research_admin():
    guard = require_admin()
    if guard is not None:
        return guard

    if not engine:
        return render_template("research_admin.html", db_ready=False, overview=[])

    overview = db_fetch_class_overview()
    return render_template("research_admin.html", db_ready=True, overview=overview)


@app.route("/research/export/student_sessions.xlsx")
def export_student_sessions_xlsx():
    guard = require_admin()
    if guard is not None:
        return guard

    # XLSX 기능은 openpyxl 의존: 미설치면 라우트만 500으로 종료
    if not OPENPYXL_AVAILABLE or Workbook is None:
        return "openpyxl not installed on server", 500

    if not engine:
        return "DB not configured", 400

    class_code = (request.args.get("class_code") or "").strip().upper()
    sid = (request.args.get("sid") or "").strip()
    if not class_code or not sid:
        return "class_code and sid are required", 400

    try:
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT class_code, sid, student_name, submitted, confidence, priority, created_at, placements
                FROM student_sessions
                WHERE class_code = :code AND sid = :sid
                ORDER BY student_name ASC
            """), {"code": class_code, "sid": sid}).fetchall()
    except Exception as e:
        return f"DB query failed: {e}", 500

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "student_sessions"

        headers = ["class_code", "sid", "student_name", "submitted", "confidence", "priority", "created_at", "placements_json"]
        ws.append(headers)

        for r in rows:
            placements_json = ""
            try:
                placements_json = json.dumps(
                    r.placements if isinstance(r.placements, dict) else (r.placements or {}),
                    ensure_ascii=False
                )
            except Exception:
                placements_json = ""

            ws.append([
                r.class_code,
                r.sid,
                r.student_name,
                bool(r.submitted),
                r.confidence,
                r.priority,
                r.created_at.isoformat() if r.created_at else None,
                placements_json,
            ])

        _autosize_columns(ws)
        return _xlsx_response(wb, f"student_sessions_{class_code}_sid{sid}.xlsx")

    except Exception as e:
        return f"xlsx generation failed: {e}", 500



@app.route("/research/export/teacher_runs.xlsx")
def export_teacher_runs_xlsx():
    guard = require_admin()
    if guard is not None:
        return guard

    if not engine:
        return "DB not configured", 400

    class_code = (request.args.get("class_code") or "").strip().upper()
    session_id = (request.args.get("session_id") or "").strip()
    if not class_code or not session_id:
        return "class_code and session_id are required", 400

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT id, class_code, session_id, teacher_username, condition, submitted,
                   started_at, ended_at, duration_ms, confidence_score, created_at, placements
            FROM teacher_placement_runs
            WHERE class_code = :code AND session_id = :sid
            ORDER BY created_at DESC
        """), {"code": class_code, "sid": session_id}).fetchall()

        run_ids = [int(r.id) for r in rows]
        decisions = []
        if run_ids:
            decisions = conn.execute(text("""
                SELECT id, run_id, target_student_name, priority_rank, decision_confidence, reason_tags, created_at
                FROM teacher_decisions
                WHERE run_id = ANY(:run_ids)
                ORDER BY run_id DESC, priority_rank ASC
            """), {"run_ids": run_ids}).fetchall()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "teacher_runs"
    ws1.append([
        "run_id", "class_code", "session_id", "teacher_username", "condition", "submitted",
        "started_at", "ended_at", "duration_ms", "confidence_score", "created_at", "placements_json"
    ])

    for r in rows:
        placements_json = ""
        try:
            placements_json = json.dumps(
                r.placements if isinstance(r.placements, dict) else (r.placements or {}),
                ensure_ascii=False
            )
        except Exception:
            placements_json = ""

        ws1.append([
            int(r.id),
            r.class_code,
            r.session_id,
            r.teacher_username,
            r.condition,
            bool(r.submitted),
            r.started_at.isoformat() if r.started_at else None,
            r.ended_at.isoformat() if r.ended_at else None,
            r.duration_ms,
            r.confidence_score,
            r.created_at.isoformat() if r.created_at else None,
            placements_json,
        ])

    _autosize_columns(ws1)

    ws2 = wb.create_sheet("teacher_decisions")
    ws2.append(["decision_id", "run_id", "target_student_name", "priority_rank", "decision_confidence", "reason_tags_json", "created_at"])

    for d in decisions:
        reason_json = ""
        try:
            reason_json = json.dumps(d.reason_tags if isinstance(d.reason_tags, (dict, list)) else (d.reason_tags or []), ensure_ascii=False)
        except Exception:
            reason_json = ""

        ws2.append([
            int(d.id),
            int(d.run_id),
            d.target_student_name,
            d.priority_rank,
            d.decision_confidence,
            reason_json,
            d.created_at.isoformat() if d.created_at else None,
        ])

    _autosize_columns(ws2)
    return _xlsx_response(wb, f"teacher_runs_{class_code}_session{session_id}.xlsx")


# -------------------------
# Teacher dashboard + class management
# -------------------------

@app.route("/teacher/dashboard")
def dashboard():
    if "teacher" not in session:
        return redirect("/teacher/login")

    classes: Dict[str, Any] = {}
    try:
        if engine:
            classes = db_list_classes_for_teacher(session["teacher"])
        else:
            d = load_data()
            classes = {c: v for c, v in d.get("classes", {}).items() if v.get("teacher") == session["teacher"]}
    except Exception:
        d = load_data()
        classes = {c: v for c, v in d.get("classes", {}).items() if v.get("teacher") == session["teacher"]}

    if classes and not session.get("selected_class"):
        session["selected_class"] = next(iter(classes.keys()))

    return render_template("dashboard.html", classes=classes)


@app.route("/teacher/create", methods=["GET", "POST"])
def create_class():
    if "teacher" not in session:
        return redirect("/teacher/login")

    if request.method == "POST":
        code = make_code()
        class_name = request.form.get("class_name", "").strip()
        students_raw = request.form.get("students", "")

        parsed: List[Dict[str, str]] = []
        auto_no = 1
        for line in students_raw.splitlines():
            line = line.strip()
            if not line:
                continue

            parts = [p.strip() for p in line.split("\t")]
            if len(parts) == 1:
                parts = [p.strip() for p in line.split(",")]

            # 허용 형식:
            # - (이름만) / (번호, 이름) / (번호, 이름, 성별)
            name = ""
            no = ""
            gender = ""

            if len(parts) == 1:
                name = parts[0]
                no = str(auto_no)
            elif len(parts) == 2:
                no, name = parts[0], parts[1]
            else:
                no, name, gender = parts[0], parts[1], (parts[2] or "").upper()

            if not name:
                continue

            if gender not in ["M", "F"]:
                gender = ""

            parsed.append({"no": str(no or auto_no), "name": name, "gender": gender})
            auto_no += 1


        if engine:
            db_create_class(
                teacher_username=session["teacher"],
                class_code=code,
                class_name=class_name or f"학급 {code}",
                students=parsed,
            )
        else:
            d = load_data()
            d.setdefault("classes", {})
            d["classes"][code] = ensure_class_schema({
                "name": class_name or f"학급 {code}",
                "teacher": session["teacher"],
                "students": parsed,
                "students_data": {s["name"]: {"sessions": {}} for s in parsed},
                "sessions": {}
            })
            save_data_safely(d)

        session["selected_class"] = code
        return redirect("/teacher/dashboard")

    return render_template("create_class.html")


@app.route("/teacher/class/delete", methods=["POST"])
def teacher_delete_class():
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (request.form.get("code") or "").upper().strip()
    if not code:
        return redirect("/teacher/dashboard")

    try:
        if engine:
            db_delete_class_for_teacher(code, session["teacher"])
        else:
            d = load_data()
            cls = d.get("classes", {}).get(code)
            if cls and cls.get("teacher") == session["teacher"]:
                d["classes"].pop(code, None)
                save_data_safely(d)
    except Exception:
        pass

    # If deleted selected class, clear
    if session.get("selected_class") == code:
        session.pop("selected_class", None)
    return redirect("/teacher/dashboard")

@app.route("/teacher/class/<code>/students", methods=["GET", "POST"])
def teacher_manage_students(code: str):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        # 1) 전입(추가)
        if action == "add":
            student_no = (request.form.get("student_no") or "").strip()
            name = (request.form.get("name") or "").strip()
            gender = (request.form.get("gender") or "").upper().strip()
            if gender not in ["M", "F"]:
                gender = ""

            if name:
                import random
                # 학급 내 pin 중복 없이 생성
                with engine.begin() as conn:
                    existing = conn.execute(text("""
                        SELECT pin_code FROM students WHERE class_code=:c
                    """), {"c": code}).fetchall()
                    used = {r.pin_code for r in existing if r.pin_code}

                    pin = None
                    for _ in range(2000):
                        cand = str(random.randint(100000, 999999))
                        if cand not in used:
                            pin = cand
                            break
                    if not pin:
                        pin = "100000"
                        while pin in used:
                            pin = str(int(pin) + 1)

                    conn.execute(text("""
                        INSERT INTO students (class_code, student_no, name, gender, pin_code, active, joined_at)
                        VALUES (:c, :no, :n, :g, :p, TRUE, NOW())
                    """), {"c": code, "no": student_no, "n": name, "g": gender, "p": pin})

        # 2) 전출/복귀(활성 토글)
        if action == "toggle_active":
            name = (request.form.get("name") or "").strip()
            make_inactive = (request.form.get("set") or "") == "0"
            if name:
                with engine.begin() as conn:
                    if make_inactive:
                        conn.execute(text("""
                            UPDATE students
                            SET active = FALSE, left_at = NOW()
                            WHERE class_code = :c AND name = :n
                        """), {"c": code, "n": name})
                    else:
                        conn.execute(text("""
                            UPDATE students
                            SET active = TRUE, left_at = NULL
                            WHERE class_code = :c AND name = :n
                        """), {"c": code, "n": name})

        # 3) 성별 수정
        if action == "set_gender":
            name = (request.form.get("name") or "").strip()
            gender = (request.form.get("gender") or "").upper().strip()
            if gender not in ["M", "F", ""]:
                gender = ""
            if name:
                with engine.begin() as conn:
                    conn.execute(text("""
                        UPDATE students
                        SET gender = :g
                        WHERE class_code = :c AND name = :n
                    """), {"c": code, "n": name, "g": gender})

        return redirect(f"/teacher/class/{code}/students")

    # GET: 학생 목록 표시
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT student_no, name, gender, pin_code, active
            FROM students
            WHERE class_code = :c
            ORDER BY id ASC
        """), {"c": code}).fetchall()

    students = []
    for r in rows:
        students.append({
            "no": r.student_no or "",
            "name": r.name,
            "gender": r.gender or "",
            "pin": r.pin_code or "",
            "active": bool(r.active),
        })

    return render_template("teacher_manage_students.html", cls=cls, code=code, students=students)



@app.route("/teacher/class/<code>")
def class_detail(code):
    # v2 is now the primary class detail page
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (request.args.get("sid") or session.get("selected_session") or "1").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"
    session["selected_session"] = sid
    session["selected_class"] = code
    return redirect(f"/teacher/class/{code}/v2?sid={sid}&open=1")


@app.route("/teacher/class/<code>/legacy")
def class_detail_legacy(code):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()

    sid = (request.args.get("sid") or session.get("selected_session") or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"
    session["selected_session"] = sid

    if engine:
        cls = db_get_class_for_teacher(code, session["teacher"])
        if not cls or cls.get("_forbidden"):
            return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

        cls = ensure_class_schema(cls)
        students = db_get_students_in_class(code)
        cls["students"] = students

        submitted_map = db_get_submitted_map(code, sid)
        session["selected_class"] = code

        rows: List[Dict[str, Any]] = []
        for i, item in enumerate(students, start=1):
            no = str(item.get("no", "") or i)
            name = (item.get("name") or "").strip()
            if not name:
                continue
            submitted = bool(submitted_map.get(name, False))
            status = "완료" if submitted else "미완료"
            rows.append({"no": no, "name": name, "status": status, "url_name": quote(name)})

        session_links: List[Dict[str, str]] = []
        for _sid, meta in sorted(cls.get("sessions", {}).items(), key=lambda x: int(x[0])):
            session_links.append({"sid": _sid, "label": meta.get("label", f"{_sid}차"), "url": f"/s/{code}/{_sid}"})

        return render_template("class_detail.html", cls=cls, code=code, rows=rows, sid=sid, session_links=session_links)

    # JSON fallback
    d = load_data()
    cls = ensure_class_schema(d.get("classes", {}).get(code))
    if not cls or cls.get("teacher") != session["teacher"]:
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    session["selected_class"] = code

    rows = []
    for i, item in enumerate(cls.get("students", []), start=1):
        if isinstance(item, dict):
            no = str(item.get("no", "") or i)
            name = (item.get("name") or "").strip()
        else:
            no = str(i)
            name = (item or "").strip()

        if not name:
            continue

        submitted = bool(cls.get("students_data", {}).get(name, {}).get("sessions", {}).get(sid, {}).get("submitted", False))
        status = "완료" if submitted else "미완료"
        rows.append({"no": no, "name": name, "status": status, "url_name": quote(name)})

    session_links = []
    for _sid, meta in sorted(cls.get("sessions", {}).items(), key=lambda x: int(x[0])):
        session_links.append({"sid": _sid, "label": meta.get("label", f"{_sid}차"), "url": f"/s/{code}/{_sid}"})

    return render_template("class_detail.html", cls=cls, code=code, rows=rows, sid=sid, session_links=session_links)

@app.route("/teacher/class/<code>/v2")
def class_detail_v2(code):
    """Parallel rebuild: new class detail UI (v2).
    Keeps existing v1 intact for safe rollout.

    Notes:
    - v2 currently supports research plan default of 4 rounds (sid 1-4).
    - v2 uses `open=1` query param to expand the selected round panel.
    """
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (request.args.get("sid") or session.get("selected_session") or "1").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"
    session["selected_session"] = sid

    open_panel = (request.args.get("open") or "").strip() == "1"

    if engine:
        cls = db_get_class_for_teacher(code, session["teacher"])
        if not cls or cls.get("_forbidden"):
            return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

        cls = ensure_class_schema(cls)
        students = db_get_students_in_class(code)
        cls["students"] = students

        submitted_map = db_get_submitted_map(code, sid)
        session["selected_class"] = code

        rows: List[Dict[str, Any]] = []
        for i, item in enumerate(students, start=1):
            no = str(item.get("no", "") or i)
            name = (item.get("name") or "").strip()
            if not name:
                continue
            submitted = bool(submitted_map.get(name, False))
            status = "완료" if submitted else "미완료"
            rows.append({"no": no, "name": name, "status": status, "url_name": quote(name)})

        session_links: List[Dict[str, str]] = []
        for _sid, meta in sorted(cls.get("sessions", {}).items(), key=lambda x: int(x[0])):
            # v2 uses 1-4 by default; still render whatever exists in schema, but UI buttons will show 1-4.
            session_links.append({"sid": _sid, "label": meta.get("label", f"{_sid}차"), "url": f"/s/{code}/{_sid}"})

        return render_template(
            "class_detail_v2.html",
            cls=cls,
            code=code,
            rows=rows,
            sid=sid,
            session_links=session_links,
            open_panel=open_panel,
        )

    # JSON fallback
    d = load_data()
    cls = ensure_class_schema(d.get("classes", {}).get(code))
    if not cls or cls.get("teacher") != session["teacher"]:
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    session["selected_class"] = code

    rows = []
    for i, item in enumerate(cls.get("students", []), start=1):
        if isinstance(item, dict):
            no = str(item.get("no", "") or i)
            name = (item.get("name") or "").strip()
        else:
            no = str(i)
            name = (item or "").strip()

        if not name:
            continue

        submitted = bool(cls.get("students_data", {}).get(name, {}).get("sessions", {}).get(sid, {}).get("submitted", False))
        status = "완료" if submitted else "미완료"
        rows.append({"no": no, "name": name, "status": status, "url_name": quote(name)})

    session_links = []
    for _sid, meta in sorted(cls.get("sessions", {}).items(), key=lambda x: int(x[0])):
        session_links.append({"sid": _sid, "label": meta.get("label", f"{_sid}차"), "url": f"/s/{code}/{_sid}"})

    return render_template(
        "class_detail_v2.html",
        cls=cls,
        code=code,
        rows=rows,
        sid=sid,
        session_links=session_links,
        open_panel=open_panel,
    )


# -------------------------
# v2 API: finalize flow
# -------------------------

def _require_teacher() -> Optional[str]:
    t = session.get("teacher")
    if not t:
        return None
    return str(t)


def _require_class_access(code: str, teacher_username: str) -> Optional[Dict[str, Any]]:
    code = (code or "").upper().strip()
    cls = db_get_class_for_teacher(code, teacher_username)
    if not cls or cls.get("_forbidden"):
        return None
    return ensure_class_schema(cls)


@app.route("/api/v2/class/<code>/session/<sid>/state")
def api_v2_state(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401

    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403

    students = db_get_students_in_class(code)
    submitted_map = db_get_submitted_map(code, sid)
    incomplete = []
    for item in students:
        name = (item.get("name") or "").strip()
        if not name:
            continue
        if not bool(submitted_map.get(name, False)):
            incomplete.append(name)

    state = db_get_session_state_v2(code, sid, t)
    state["ok"] = True
    state["incomplete_students"] = incomplete
    return jsonify(state)


@app.route("/api/v2/class/<code>/session/<sid>/preview_seen", methods=["POST"])
def api_v2_preview_seen(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401
    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403

    db_set_preview_seen_v2(code, sid, t)
    return jsonify({"ok": True, "preview_seen": True})


@app.route("/api/v2/class/<code>/session/<sid>/exclusions", methods=["POST"])
def api_v2_exclusions(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401
    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403

    payload = request.get_json(silent=True) or {}
    mode = (payload.get("mode") or "").strip()
    items = payload.get("items") or []
    if mode not in ["no_exclude", "exclude"]:
        return jsonify({"ok": False, "error": "BAD_REQUEST"}), 400
    if mode == "exclude":
        # validate reasons present
        norm_items = []
        for it in items:
            if not isinstance(it, dict):
                continue
            nm = (it.get("student_name") or "").strip()
            rs = (it.get("reason") or "").strip()
            if not nm:
                continue
            if not rs:
                return jsonify({"ok": False, "error": "REASON_REQUIRED"}), 400
            norm_items.append({"student_name": nm, "reason": rs})
        items = norm_items

    db_set_exclusions_v2(code, sid, t, mode, items)
    return jsonify({"ok": True, "exclusions_resolved": True})


@app.route("/api/v2/class/<code>/session/<sid>/survey", methods=["POST"])
def api_v2_survey(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401
    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403

    payload = request.get_json(silent=True) or {}
    db_upsert_survey_v2(code, sid, t, payload)
    return jsonify({"ok": True, "survey_submitted": True})


@app.route("/api/v2/class/<code>/session/<sid>/finalize", methods=["POST"])
def api_v2_finalize(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401
    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403

    res = db_finalize_session_v2(code, sid, t)
    if not res.get("ok"):
        return jsonify(res), 400

    return jsonify({
        "ok": True,
        "finalized": True,
        "result_url": f"/teacher/class/{code}/result/{sid}"
    })


# -------------------------
# v2: session-level result page (minimal)
# -------------------------
@app.route("/teacher/class/<code>/result/<sid>")
def teacher_session_result(code: str, sid: str):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404
    cls = ensure_class_schema(cls)

    students = db_get_students_in_class(code)
    submitted_map = db_get_submitted_map(code, sid)
    done = 0
    total = 0
    for it in students:
        nm = (it.get("name") or "").strip()
        if not nm:
            continue
        total += 1
        if bool(submitted_map.get(nm, False)):
            done += 1

    # teacher placement existence
    teacher_has = False
    if engine:
        with engine.connect() as conn:
            r = conn.execute(text("""
                SELECT 1 FROM teacher_placement_runs
                WHERE class_code=:code AND teacher_username=:t AND session_id=:sid
                LIMIT 1
            """), {"code": code, "t": session["teacher"], "sid": sid}).fetchone()
            teacher_has = bool(r)

    state = db_get_session_state_v2(code, sid, session["teacher"])
    return render_template(
        "session_result.html",
        cls=cls,
        code=code,
        sid=sid,
        total=total,
        done=done,
        teacher_has=teacher_has,
        state=state,
    )

@app.route("/teacher/class/<code>/result/<sid>/<url_name>")
def teacher_view_student_result(code, sid, url_name):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    student_name = (unquote(url_name) or "").strip()

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    students = db_get_students_in_class(code)
    all_names = [s["name"] for s in students]
    if student_name not in all_names:
        return "학생을 찾을 수 없습니다.", 404

    friends = [n for n in all_names if n != student_name]
    sess = db_get_student_session(code, student_name, sid)
    placements = (sess.get("placements") if sess else {}) or {}

    student_session = {"placements": placements, "submitted": True}

    cls_for_view = ensure_class_schema({
        "code": code,
        "name": db_get_class_name(code) or code,
        "teacher": session["teacher"],
        "sessions": {}
    }) or {}
    session_meta = (cls_for_view.get("sessions") or {}).get(sid, {"label": f"{sid}차"})

    return render_template(
        "student_write.html",
        name=student_name,
        friends=friends,
        placements=placements,
        student_session=student_session,
        sid=sid,
        session_meta=session_meta,
        teacher_view=True,
    )


# -------------------------
# Student entry
# -------------------------

@app.route("/s/<code>/<sid>", methods=["GET", "POST"])
def student_enter_session(code, sid):
    code = (code or "").upper().strip()
    sid = (sid or "1").strip()

    if engine:
        with engine.connect() as conn:
            row = conn.execute(text("""
                SELECT code, name, teacher_username
                FROM classes
                WHERE code = :code
                LIMIT 1
            """), {"code": code}).fetchone()

        if not row:
            return "학급을 찾을 수 없습니다.", 404

        cls = ensure_class_schema({"code": code, "name": row.name, "teacher": row.teacher_username, "sessions": {}})
        students = db_get_students_in_class(code)
        cls["students"] = students
        cls["students_data"] = {s["name"]: {"sessions": {}} for s in students}
    else:
        d = load_data()
        cls = ensure_class_schema(d.get("classes", {}).get(code))
        if not cls:
            return "학급을 찾을 수 없습니다.", 404

    if sid not in (cls.get("sessions") or {}):
        sid = "1"

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if name not in (cls.get("students_data") or {}):
            return render_template(
                "student_enter_session.html",
                error="학생 명단에 없는 이름입니다.",
                code=code,
                sid=sid,
                session_label=(cls.get("sessions") or {}).get(sid, {}).get("label", f"{sid}차"),
            )

        session["code"] = code
        session["name"] = name
        session["sid"] = sid
        session["selected_class"] = code
        session["selected_session"] = sid

        return redirect("/student/write")

    return render_template(
        "student_enter_session.html",
        code=code,
        sid=sid,
        session_label=(cls.get("sessions") or {}).get(sid, {}).get("label", f"{sid}차"),
    )


@app.route("/qr/<code>/<sid>.png")
def qr_session_link(code, sid):
    code = (code or "").upper().strip()
    sid = str(sid).strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    if engine:
        with engine.connect() as conn:
            row = conn.execute(text("SELECT 1 FROM classes WHERE code = :code LIMIT 1"), {"code": code}).fetchone()
        if not row:
            return "학급을 찾을 수 없습니다.", 404
    else:
        d = load_data()
        cls = d.get("classes", {}).get(code)
        if not cls:
            return "학급을 찾을 수 없습니다.", 404

    base = request.url_root.rstrip("/")
    target = f"{base}/s/{code}/{sid}"

    try:
        import qrcode
    except ModuleNotFoundError:
        return "QR 코드 생성을 위해 qrcode 라이브러리가 필요합니다.", 500

    img = qrcode.make(target)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)

    resp = send_file(buf, mimetype="image/png")
    resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp


# Legacy student enter (optional)
@app.route("/student", methods=["GET", "POST"])
def student_enter():
    # Keep legacy entry route as a fallback to sid=1
    if request.method == "POST":
        code = request.form.get("code", "").upper().strip()
        name = request.form.get("name", "").strip()

        if engine:
            with engine.connect() as conn:
                c_row = conn.execute(text("SELECT 1 FROM classes WHERE code=:code LIMIT 1"), {"code": code}).fetchone()
            if not c_row:
                return render_template("student_enter.html", error="입장 실패")

            pin = (request.form.get("pin") or "").strip()

            # 6자리 숫자만 허용(아이들 입력 실수 방지)
            if not (len(pin) == 6 and pin.isdigit()):
                return render_template("student_enter.html", error="개인 코드는 6자리 숫자여야 합니다.")

            with engine.connect() as conn:
                s_row = conn.execute(text("""
                    SELECT 1
                    FROM students
                    WHERE class_code = :code
                      AND name = :name
                      AND pin_code = :pin
                      AND active = TRUE
                    LIMIT 1
                """), {"code": code, "name": name, "pin": pin}).fetchone()

            if not s_row:
                return render_template("student_enter.html", error="입장 실패")

            session["code"] = code
            session["name"] = name
            session["sid"] = "1"
            session["selected_class"] = code
            session["selected_session"] = "1"
            return redirect("/student/write")

        d = load_data()
        cls = d.get("classes", {}).get(code)
        if not cls or name not in (cls.get("students_data") or {}):
            return render_template("student_enter.html", error="입장 실패")

        session["code"] = code
        session["name"] = name
        session["sid"] = "1"
        session["selected_class"] = code
        session["selected_session"] = "1"
        return redirect("/student/write")

    return render_template("student_enter.html")


# -------------------------
# Student write
# -------------------------

@app.route("/student/write", methods=["GET", "POST"])
def student_write():
    if "code" not in session or "name" not in session:
        return redirect("/student")

    code = (session.get("code") or "").upper().strip()
    name = (session.get("name") or "").strip()

    sid = (session.get("sid") or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"
        session["sid"] = sid

    if engine:
        with engine.connect() as conn:
            row = conn.execute(text("""
                SELECT code, name, teacher_username
                FROM classes
                WHERE code = :code
                LIMIT 1
            """), {"code": code}).fetchone()
        if not row:
            return redirect("/student")

        students = db_get_students_in_class(code)
        all_names = [s["name"] for s in students]
        if name not in all_names:
            return redirect("/student")

        friends = [n for n in all_names if n != name]

        cls = ensure_class_schema({"code": code, "name": row.name, "teacher": row.teacher_username, "sessions": {}})
        cls["students"] = students
        cls["students_data"] = {n: {"sessions": {}} for n in all_names}

        db_sess = db_get_student_session(code, name, sid)
        placements = (db_sess.get("placements") if db_sess else {}) or {}
        submitted = bool(db_sess.get("submitted")) if db_sess else False

        if request.method == "POST":
            if submitted:
                return redirect("/student/submitted")

            placements_json = (request.form.get("placements_json") or "{}").strip()
            try:
                placements_obj = json.loads(placements_json) if placements_json else {}
            except Exception:
                placements_obj = {}

            resp = post_to_sheet({
                "action": "result_append",
                "teacher": row.teacher_username,
                "class_code": code,
                "student": name,
                "session": sid,
                "placements": placements_obj,
                "ip": request.headers.get("X-Forwarded-For", request.remote_addr) or "",
            })

            if resp.get("status") != "ok":
                return render_template(
                    "student_write.html",
                    error=f"저장 실패(구글 시트): {resp}",
                    name=name,
                    friends=friends,
                    placements=placements_obj,
                    student_session={"placements": placements_obj, "submitted": False},
                    sid=sid,
                    session_meta=(cls.get("sessions") or {}).get(sid, {}),
                )

            db_upsert_student_session(code, name, sid, placements_obj, submitted=True)
            return redirect("/student/submitted")

        return render_template(
            "student_write.html",
            name=name,
            friends=friends,
            placements=placements,
            student_session={"placements": placements, "submitted": submitted},
            sid=sid,
            session_meta=(cls.get("sessions") or {}).get(sid, {}),
        )

    # JSON fallback
    d = load_data()
    cls = ensure_class_schema(d.get("classes", {}).get(code))
    if not cls:
        return redirect("/student")
    if name not in (cls.get("students_data") or {}):
        return redirect("/student")
    if sid not in (cls.get("sessions") or {}):
        sid = "1"
        session["sid"] = sid

    student = cls["students_data"][name]
    student.setdefault("sessions", {})
    student["sessions"].setdefault(sid, {"placements": {}, "submitted": False})
    ssession = student["sessions"][sid]

    friends = [s["name"] for s in cls.get("students", []) if isinstance(s, dict) and s.get("name") != name]
    placements = ssession.get("placements") or {}

    if request.method == "POST":
        if ssession.get("submitted"):
            return redirect("/student/submitted")

        placements_json = (request.form.get("placements_json") or "{}").strip()
        try:
            placements_obj = json.loads(placements_json) if placements_json else {}
        except Exception:
            placements_obj = {}

        resp = post_to_sheet({
            "action": "result_append",
            "teacher": cls.get("teacher", ""),
            "class_code": code,
            "student": name,
            "session": sid,
            "placements": placements_obj,
            "ip": request.headers.get("X-Forwarded-For", request.remote_addr) or "",
        })

        if resp.get("status") != "ok":
            return render_template(
                "student_write.html",
                error=f"저장 실패(구글 시트): {resp}",
                name=name,
                friends=friends,
                placements=placements_obj,
                student_session=ssession,
                sid=sid,
                session_meta=(cls.get("sessions") or {}).get(sid, {}),
            )

        ssession["placements"] = placements_obj
        ssession["submitted"] = True
        student["sessions"][sid] = ssession
        d["classes"][code] = ensure_class_schema(cls)
        save_data_safely(d)
        return redirect("/student/submitted")

    return render_template(
        "student_write.html",
        name=name,
        friends=friends,
        placements=placements,
        student_session=ssession,
        sid=sid,
        session_meta=(cls.get("sessions") or {}).get(sid, {}),
    )


@app.route("/student/submitted")
def student_submitted():
    return render_template("student_submitted.html")


# -------------------------
# Teacher placement flow
# -------------------------

@app.route("/teacher/class/<code>/placement/start")
def teacher_placement_start(code):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (request.args.get("sid") or session.get("selected_session") or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    condition = (request.args.get("condition") or "BASELINE").strip()
    if condition not in ["BASELINE", "TOOL_ASSISTED"]:
        condition = "BASELINE"

    tool_run_id = request.args.get("tool_run_id")
    tool_run_id_val = int(tool_run_id) if (tool_run_id and str(tool_run_id).isdigit()) else None

    run_id = db_create_teacher_run(code, session["teacher"], sid, condition, tool_run_id=tool_run_id_val)
    return redirect(f"/teacher/placement/{run_id}")


@app.route("/teacher/placement/<int:run_id>", methods=["GET", "POST"])
def teacher_placement_write(run_id: int):
    if "teacher" not in session:
        return redirect("/teacher/login")

    run = db_get_teacher_run(run_id)
    if not run or run["teacher_username"] != session["teacher"]:
        return "접근 권한이 없습니다.", 403

    code = run["class_code"]
    sid = run["session_id"]

    students = db_get_students_in_class(code)
    all_names = [s["name"] for s in students]
    placements = run.get("placements") or {}

    if request.method == "POST":
        placements_json = (request.form.get("placements_json") or "{}").strip()
        try:
            placements_obj = json.loads(placements_json) if placements_json else {}
        except Exception:
            placements_obj = {}

        db_update_teacher_run_placements(run_id, placements_obj)
        return redirect(f"/teacher/placement/{run_id}/complete")

    return render_template(
        "teacher_write.html",
        run=run,
        code=code,
        sid=sid,
        friends=all_names,
        placements=placements,
    )

@app.route("/teacher/placement/<int:run_id>/complete", methods=["GET", "POST"])
def teacher_placement_complete(run_id: int):
    if "teacher" not in session:
        return redirect("/teacher/login")

    run = db_get_teacher_run(run_id)
    if not run or run["teacher_username"] != session["teacher"]:
        return "접근 권한이 없습니다.", 403

    # 완료 화면에서 자동완성 목록 제공용
    students = db_get_students_in_class(run["class_code"]) if engine else []

    if request.method == "POST":
        duration_ms_raw = request.form.get("duration_ms") or "0"
        confidence_raw = (request.form.get("confidence_score") or "").strip()

        try:
            duration_ms = int(duration_ms_raw)
        except Exception:
            duration_ms = 0

        # 슬라이더: 0~100 int (범위 밖이면 자동 보정)
        try:
            confidence_score = int(confidence_raw) if confidence_raw != "" else 50
        except Exception:
            confidence_score = 50
        confidence_score = max(0, min(100, confidence_score))

        # 우선순위는 hidden input(priority_1~3)에 실려 옴
        decisions: List[Dict[str, Any]] = []
        for rank in [1, 2, 3]:
            nm = (request.form.get(f"priority_{rank}") or "").strip()
            if nm:
                decisions.append({"name": nm, "rank": rank})

        db_replace_teacher_decisions(run_id, decisions)
        db_complete_teacher_run(run_id, duration_ms=duration_ms, confidence_score=confidence_score)

        return redirect(f"/teacher/class/{run['class_code']}?sid={run['session_id']}")


    return render_template("teacher_complete.html", run=run, students=students)



# -------------------------
# Analysis helpers
# -------------------------

def _extract_point(v: Any, canvas_w: Optional[float] = None, canvas_h: Optional[float] = None) -> Optional[Tuple[float, float, str]]:
    """
    placements item v -> (x,y,mode_tag)
    - abs: {x,y,w,h,mode:'abs'} normalized by w/h if present
    - rel: {x,y,...} used as-is, later bbox-normalized
    """
    if not isinstance(v, dict):
        return None

    mode = v.get("mode")
    x = v.get("x")
    y = v.get("y")
    if not (isinstance(x, (int, float)) and isinstance(y, (int, float))):
        return None

    if mode == "abs":
        w = v.get("w") if isinstance(v.get("w"), (int, float)) else canvas_w
        h = v.get("h") if isinstance(v.get("h"), (int, float)) else canvas_h
        if not (isinstance(w, (int, float)) and w > 0 and isinstance(h, (int, float)) and h > 0):
            return (float(x), float(y), "abs_raw")
        return (float(x) / float(w), float(y) / float(h), "abs_norm")

    return (float(x), float(y), "rel")


def points_from_placements_all_students(placements: Dict[str, Any], names: List[str]) -> Tuple[List[Tuple[float, float]], List[bool]]:
    pts: List[Tuple[float, float]] = []
    valid: List[bool] = []
    for nm in names:
        p = _extract_point(placements.get(nm))
        if p is None:
            pts.append((0.0, 0.0))
            valid.append(False)
        else:
            pts.append((p[0], p[1]))
            valid.append(True)

    xs = [pts[i][0] for i in range(len(pts)) if valid[i]]
    ys = [pts[i][1] for i in range(len(pts)) if valid[i]]
    if len(xs) >= 2 and len(ys) >= 2:
        minx, maxx = min(xs), max(xs)
        miny, maxy = min(ys), max(ys)
        rx = (maxx - minx) if (maxx - minx) > 1e-9 else 1.0
        ry = (maxy - miny) if (maxy - miny) > 1e-9 else 1.0
        pts = [((x - minx) / rx, (y - miny) / ry) for (x, y) in pts]

    return pts, valid


def points_from_student_session(placements: Dict[str, Any], names: List[str], self_name: str) -> Tuple[List[Tuple[float, float]], List[bool]]:
    """
    Student: self_name is included as (0,0) and valid=True.
    Others: from placements; missing -> invalid.
    Then bbox-normalize based on valid points.
    """
    pts: List[Tuple[float, float]] = []
    valid: List[bool] = []

    for nm in names:
        if nm == self_name:
            pts.append((0.0, 0.0))
            valid.append(True)
            continue

        v = placements.get(nm)
        if not isinstance(v, dict):
            pts.append((0.0, 0.0))
            valid.append(False)
            continue

        x = v.get("x")
        y = v.get("y")
        if not (isinstance(x, (int, float)) and isinstance(y, (int, float))):
            pts.append((0.0, 0.0))
            valid.append(False)
            continue

        pts.append((float(x), float(y)))
        valid.append(True)

    xs = [pts[i][0] for i in range(len(pts)) if valid[i]]
    ys = [pts[i][1] for i in range(len(pts)) if valid[i]]
    if len(xs) >= 2 and len(ys) >= 2:
        minx, maxx = min(xs), max(xs)
        miny, maxy = min(ys), max(ys)
        rx = (maxx - minx) if (maxx - minx) > 1e-9 else 1.0
        ry = (maxy - miny) if (maxy - miny) > 1e-9 else 1.0
        pts = [((x - minx) / rx, (y - miny) / ry) for (x, y) in pts]

    return pts, valid


def distance_matrix(points: List[Tuple[float, float]], valid: List[bool]) -> List[List[Optional[float]]]:
    n = len(points)
    D: List[List[Optional[float]]] = [[None] * n for _ in range(n)]
    for i in range(n):
        if not valid[i]:
            continue
        D[i][i] = 0.0
        xi, yi = points[i]
        for j in range(i + 1, n):
            if not valid[j]:
                continue
            xj, yj = points[j]
            d = math.hypot(xi - xj, yi - yj)
            d = round(float(d), 6)
            D[i][j] = d
            D[j][i] = d
    return D


def mean_distance_matrix(mats: List[List[List[Optional[float]]]]) -> List[List[Optional[float]]]:
    if not mats:
        return []
    n = len(mats[0])
    avg: List[List[Optional[float]]] = [[None] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            vals: List[float] = []
            for M in mats:
                v = M[i][j]
                if isinstance(v, (int, float)):
                    vals.append(float(v))
            if vals:
                avg[i][j] = round(sum(vals) / len(vals), 6)
    for i in range(n):
        avg[i][i] = 0.0
    return avg


def classical_mds_2d(D: List[List[Optional[float]]]) -> List[Tuple[float, float]]:
    n = len(D)
    if n == 0:
        return []

    filled: List[List[float]] = [[0.0] * n for _ in range(n)]
    for i in range(n):
        row_vals = [D[i][j] for j in range(n) if isinstance(D[i][j], (int, float)) and i != j]
        row_mean = (sum(row_vals) / len(row_vals)) if row_vals else 0.0
        for j in range(n):
            v = D[i][j]
            filled[i][j] = float(v) if isinstance(v, (int, float)) else float(row_mean)

    D2 = [[filled[i][j] ** 2 for j in range(n)] for i in range(n)]
    row_mean = [sum(D2[i]) / n for i in range(n)]
    col_mean = [sum(D2[i][j] for i in range(n)) / n for j in range(n)]
    total_mean = sum(row_mean) / n

    B = [[-0.5 * (D2[i][j] - row_mean[i] - col_mean[j] + total_mean) for j in range(n)] for i in range(n)]

    def matvec(M: List[List[float]], v: List[float]) -> List[float]:
        return [sum(M[i][k] * v[k] for k in range(n)) for i in range(n)]

    def dot(a: List[float], b: List[float]) -> float:
        return sum(a[i] * b[i] for i in range(n))

    def norm(v: List[float]) -> float:
        return math.sqrt(dot(v, v)) + 1e-12

    def power_iter(M: List[List[float]], iters: int = 80) -> Tuple[float, List[float]]:
        v = [1.0 / math.sqrt(n)] * n
        for _ in range(iters):
            w = matvec(M, v)
            nv = norm(w)
            v = [x / nv for x in w]
        lam = dot(v, matvec(M, v))
        return lam, v

    lam1, v1 = power_iter(B)
    B2 = [[B[i][j] - lam1 * v1[i] * v1[j] for j in range(n)] for i in range(n)]
    lam2, v2 = power_iter(B2)

    lam1 = max(lam1, 0.0)
    lam2 = max(lam2, 0.0)
    s1 = math.sqrt(lam1)
    s2 = math.sqrt(lam2)

    coords = [(round(v1[i] * s1, 6), round(v2[i] * s2, 6)) for i in range(n)]
    return coords


def teacher_run_distance_payload(class_code: str, sid: str, run_id: int) -> Optional[Dict[str, Any]]:
    run = db_get_teacher_run(run_id)
    if not run:
        return None

    students = db_get_students_in_class(class_code)
    names = [s["name"] for s in students]
    placements = run.get("placements") or {}

    pts, valid = points_from_placements_all_students(placements, names)
    D = distance_matrix(pts, valid)
    X = classical_mds_2d(D)

    return {
        "class_code": class_code,
        "session_id": sid,
        "run_id": run_id,
        "names": names,
        "points_norm": [{"x": pts[i][0], "y": pts[i][1], "valid": bool(valid[i])} for i in range(len(names))],
        "distance_matrix": D,
        "mds_2d": [{"x": X[i][0], "y": X[i][1]} for i in range(len(names))],
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


def student_avg_distance_payload(class_code: str, sid: str) -> Dict[str, Any]:
    students = db_get_students_in_class(class_code)
    names = [s["name"] for s in students]
    n_total = len(names)

    if n_total == 0:
        return {
            "class_code": class_code,
            "session_id": sid,
            "names": [],
            "n_total": 0,
            "n_submitted": 0,
            "submitted_students": [],
            "avg_distance_matrix": [],
            "mds_2d": [],
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    submitted = db_list_submitted_student_sessions(class_code, sid)

    mats: List[List[List[Optional[float]]]] = []
    used_students: List[str] = []

    for item in submitted:
        self_name = item.get("student_name")
        placements = item.get("placements") or {}

        if self_name not in names:
            continue

        pts, valid = points_from_student_session(placements, names, self_name=self_name)
        D = distance_matrix(pts, valid)
        mats.append(D)
        used_students.append(self_name)

    avgD = mean_distance_matrix(mats)
    X = classical_mds_2d(avgD) if avgD else []

    return {
        "class_code": class_code,
        "session_id": sid,
        "names": names,
        "n_total": n_total,
        "n_submitted": len(used_students),
        "submitted_students": used_students,
        "avg_distance_matrix": avgD,
        "mds_2d": [{"x": X[i][0], "y": X[i][1]} for i in range(len(X))],
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


def student_vs_avg_distance_payload(class_code: str, sid: str, student_name: str) -> Optional[Dict[str, Any]]:
    students = db_get_students_in_class(class_code)
    names = [s["name"] for s in students]
    if not names or student_name not in names:
        return None

    avg_cache_key = f"student_avg_{sid}"
    avg_payload = cache_get(class_code, sid, avg_cache_key)
    if not avg_payload:
        avg_payload = student_avg_distance_payload(class_code, sid)

    avgD = avg_payload.get("avg_distance_matrix") or []
    if not avgD:
        return {
            "class_code": class_code,
            "session_id": sid,
            "student_name": student_name,
            "error": "avg_distance_matrix empty",
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    ss = db_get_student_session(class_code, student_name, sid)
    if not ss:
        return None
    if not ss.get("submitted"):
        return {
            "class_code": class_code,
            "session_id": sid,
            "student_name": student_name,
            "error": "student session not submitted",
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    placements = ss.get("placements") or {}
    pts, valid = points_from_student_session(placements, names, self_name=student_name)
    Ds = distance_matrix(pts, valid)

    n = len(names)
    diffs: List[float] = []
    abs_diffs: List[float] = []

    used_pairs = 0
    total_pairs = 0

    for i in range(n):
        for j in range(i + 1, n):
            total_pairs += 1
            v_s = Ds[i][j]
            v_a = avgD[i][j] if (i < len(avgD) and j < len(avgD[i])) else None
            if isinstance(v_s, (int, float)) and isinstance(v_a, (int, float)):
                d = float(v_s) - float(v_a)
                diffs.append(d)
                abs_diffs.append(abs(d))
                used_pairs += 1

    if used_pairs == 0:
        return {
            "class_code": class_code,
            "session_id": sid,
            "student_name": student_name,
            "n_total_students": n,
            "n_pairs_total": total_pairs,
            "n_pairs_used": 0,
            "error": "no comparable pairs",
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    mean_abs = sum(abs_diffs) / len(abs_diffs)
    mean_signed = sum(diffs) / len(diffs)
    var_abs = sum((x - mean_abs) ** 2 for x in abs_diffs) / len(abs_diffs)

    self_idx = names.index(student_name)
    self_peer_diffs: List[Dict[str, Any]] = []
    for j in range(n):
        if j == self_idx:
            continue
        v_s = Ds[self_idx][j]
        v_a = avgD[self_idx][j] if (self_idx < len(avgD) and j < len(avgD[self_idx])) else None
        if isinstance(v_s, (int, float)) and isinstance(v_a, (int, float)):
            self_peer_diffs.append({
                "peer": names[j],
                "student_dist": round(float(v_s), 6),
                "avg_dist": round(float(v_a), 6),
                "diff": round(float(v_s) - float(v_a), 6),
                "abs_diff": round(abs(float(v_s) - float(v_a)), 6),
            })
    self_peer_diffs.sort(key=lambda x: x["abs_diff"], reverse=True)

    return {
        "class_code": class_code,
        "session_id": sid,
        "student_name": student_name,
        "n_total_students": n,
        "n_pairs_total": total_pairs,
        "n_pairs_used": used_pairs,
        "mean_abs_diff": round(mean_abs, 6),
        "var_abs_diff": round(var_abs, 6),
        "mean_signed_diff": round(mean_signed, 6),
        "self_peer_diffs_top": self_peer_diffs[:10],
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


def kmeans_2d(points: List[Tuple[float, float]], k: int, n_init: int = 10, max_iter: int = 60, seed: int = 42) -> Tuple[List[int], List[Tuple[float, float]], float]:
    if not points:
        return [], [], 0.0

    n = len(points)
    k = max(1, min(int(k), n))

    def dist2(a: Tuple[float, float], b: Tuple[float, float]) -> float:
        dx = a[0] - b[0]
        dy = a[1] - b[1]
        return dx * dx + dy * dy

    best_labels: List[int] = []
    best_centers: List[Tuple[float, float]] = []
    best_inertia: Optional[float] = None

    for t in range(max(1, int(n_init))):
        rnd = random.Random(int(seed) + t)

        idx = list(range(n))
        rnd.shuffle(idx)
        idx = idx[:k]
        centers = [points[i] for i in idx]
        labels = [0] * n

        for _ in range(int(max_iter)):
            changed = False

            for i in range(n):
                p = points[i]
                best_c = 0
                best_d = dist2(p, centers[0])
                for ci in range(1, k):
                    d = dist2(p, centers[ci])
                    if d < best_d:
                        best_d = d
                        best_c = ci
                if labels[i] != best_c:
                    labels[i] = best_c
                    changed = True

            new_centers: List[Tuple[float, float]] = []
            for ci in range(k):
                members = [points[i] for i in range(n) if labels[i] == ci]
                if not members:
                    new_centers.append(points[rnd.randrange(0, n)])
                    continue
                mx = sum(p[0] for p in members) / len(members)
                my = sum(p[1] for p in members) / len(members)
                new_centers.append((mx, my))

            centers = new_centers
            if not changed:
                break

        inertia = 0.0
        for i in range(n):
            inertia += dist2(points[i], centers[labels[i]])

        if best_inertia is None or inertia < best_inertia:
            best_inertia = inertia
            best_labels = labels[:]
            best_centers = centers[:]

    return best_labels, best_centers, float(best_inertia if best_inertia is not None else 0.0)


def kmeans_summary_payload(class_code: str, sid: str, k: int) -> Dict[str, Any]:
    avg_cache_key = f"student_avg_{sid}"
    avg_payload = cache_get(class_code, sid, avg_cache_key)
    if not avg_payload:
        avg_payload = student_avg_distance_payload(class_code, sid)

    names = avg_payload.get("names") or []
    pts = avg_payload.get("mds_2d") or []
    if not names or not pts or len(names) != len(pts):
        return {
            "class_code": class_code,
            "session_id": sid,
            "k": int(k),
            "error": "avg mds_2d not available",
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    points: List[Tuple[float, float]] = []
    for p in pts:
        x = p.get("x") if isinstance(p, dict) else None
        y = p.get("y") if isinstance(p, dict) else None
        if not (isinstance(x, (int, float)) and isinstance(y, (int, float))):
            x, y = 0.0, 0.0
        points.append((float(x), float(y)))

    kk = int(k)
    if kk < 2:
        kk = 2
    if kk > 4:
        kk = 4
    if kk > len(points):
        kk = max(2, min(4, len(points)))

    labels, centers, inertia = kmeans_2d(points, k=kk, n_init=10, max_iter=60, seed=42)

    cluster_sizes = [0] * kk
    for lb in labels:
        if 0 <= int(lb) < kk:
            cluster_sizes[int(lb)] += 1

    cluster_mean_radius: List[Optional[float]] = []
    for ci in range(kk):
        members = [i for i in range(len(points)) if labels[i] == ci]
        if not members:
            cluster_mean_radius.append(None)
            continue
        cx, cy = centers[ci]
        ds = [math.hypot(points[i][0] - cx, points[i][1] - cy) for i in members]
        cluster_mean_radius.append(round(sum(ds) / len(ds), 6))

    labeled_points: List[Dict[str, Any]] = []
    for i in range(len(names)):
        labeled_points.append({
            "name": names[i],
            "x": round(points[i][0], 6),
            "y": round(points[i][1], 6),
            "cluster_id": int(labels[i]) if i < len(labels) else 0,
        })

    return {
        "class_code": class_code,
        "session_id": sid,
        "k": kk,
        "n_points": len(points),
        "cluster_sizes": cluster_sizes,
        "cluster_mean_radius": cluster_mean_radius,
        "inertia": round(float(inertia), 6),
        "centers": [{"x": round(c[0], 6), "y": round(c[1], 6)} for c in centers],
        "points": labeled_points,
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


# -------------------------
# Analysis routes
# -------------------------

@app.route("/analysis/class/<code>/<sid>/teacher_run/<int:run_id>.json")
def analysis_teacher_run(code, sid, run_id):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()

    run = db_get_teacher_run(run_id)
    if (not run) or run["class_code"] != code or run["session_id"] != sid:
        return jsonify({"error": "run not found"}), 404
    if run["teacher_username"] != session["teacher"]:
        return jsonify({"error": "forbidden"}), 403

    cache_key = f"teacher_run_{run_id}"
    cached = cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    payload = teacher_run_distance_payload(code, sid, run_id)
    if not payload:
        return jsonify({"error": "failed"}), 500

    cache_set(code, sid, cache_key, payload)
    return jsonify(payload)


@app.route("/analysis/class/<code>/<sid>/student_avg.json")
def analysis_student_avg(code, sid):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"error": "forbidden"}), 403

    cache_key = f"student_avg_{sid}"
    cached = cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    payload = student_avg_distance_payload(code, sid)
    cache_set(code, sid, cache_key, payload)
    return jsonify(payload)


@app.route("/analysis/class/<code>/<sid>/student/<path:student_name>/vs_avg.json")
def analysis_student_vs_avg(code, sid, student_name):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    student_name = unquote(student_name or "").strip()

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"error": "forbidden"}), 403

    if not student_name:
        return jsonify({"error": "student_name required"}), 400

    cache_key = f"student_vs_avg_{sid}_{student_name}"
    cached = cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    payload = student_vs_avg_distance_payload(code, sid, student_name)
    if not payload:
        return jsonify({"error": "not found"}), 404

    cache_set(code, sid, cache_key, payload)
    return jsonify(payload)


@app.route("/analysis/class/<code>/<sid>/kmeans_summary.json")
def analysis_kmeans_summary(code, sid):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"error": "forbidden"}), 403

    k_raw = request.args.get("k", "3")
    try:
        k = int(k_raw)
    except Exception:
        k = 3
    if k < 2:
        k = 2
    if k > 4:
        k = 4

    cache_key = f"kmeans_summary_{sid}_k{k}"
    cached = cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    payload = kmeans_summary_payload(code, sid, k)
    cache_set(code, sid, cache_key, payload)
    return jsonify(payload)


# -------------------------
# Main
# -------------------------

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port)
